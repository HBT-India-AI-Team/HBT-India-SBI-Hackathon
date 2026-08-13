"""Parses a Probe42/Tofler-style MCA company due-diligence report (.xls/.xlsx,
the multi-sheet "About the Company" / "GST" / "EPFO Establishments" /
"Standalone Financial Data" / "Open Charges Sequence" layout) into the flat
evidence dict fin_health's gates/factors actually score against.

Deliberately targeted at this one real report template, not a generic
"any spreadsheet" ingester — the sheet names and layouts below are read
directly off a real sample report. A different due-diligence vendor's
export would need its own parser, following this same shape: read cells,
never invent a value that isn't actually in the sheet.
"""
from __future__ import annotations

import io
from typing import Any

import openpyxl
import xlrd


def _read_rows(file_bytes: bytes, filename: str) -> dict[str, list[list[Any]]]:
    """Every sheet as a list of rows (each a list of cell values), keyed by
    sheet name — the one format-specific step; everything after this walks
    plain Python lists, not workbook objects.
    """
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        return {
            name: [[cell for cell in row] for row in wb[name].iter_rows(values_only=True)]
            for name in wb.sheetnames
        }
    if lower.endswith(".xls"):
        wb = xlrd.open_workbook(file_contents=file_bytes)
        return {
            name: [
                [wb.sheet_by_name(name).cell_value(r, c) for c in range(wb.sheet_by_name(name).ncols)]
                for r in range(wb.sheet_by_name(name).nrows)
            ]
            for name in wb.sheet_names()
        }
    raise ValueError(f"Unsupported file type: '{filename}' — expected .xls or .xlsx")


def _property_value_dict(rows: list[list[Any]]) -> dict[str, Any]:
    """"About the Company"-shaped sheets: every row is [property, value]."""
    out: dict[str, Any] = {}
    for row in rows:
        if len(row) >= 2 and row[0] not in (None, ""):
            out[str(row[0]).strip()] = row[1]
    return out


def _first_data_row_dict(rows: list[list[Any]]) -> dict[str, Any]:
    """Tabular sheets (GST, EPFO Establishments): row 0 is the header, row 1
    is the first real record — the one this report's own entity concerns,
    since these reports are always scoped to a single company.
    """
    if len(rows) < 2:
        return {}
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    return dict(zip(header, rows[1]))


def _latest_year_ratio(rows: list[list[Any]], section_label: str, metric_label: str) -> float | None:
    """"Standalone Financial Data"'s RATIOS block: rows shaped
    [metric_name, '', year1_value, year2_value] following a section header
    row whose first cell is `section_label`. Returns the latest (second)
    year's value for `metric_label`, or None if the report doesn't have it.
    """
    in_section = False
    for row in rows:
        label = str(row[0]).strip() if row and row[0] not in (None, "") else ""
        if label == section_label:
            in_section = True
            continue
        if in_section:
            if label == "":
                break  # section ends at the next blank row
            if label == metric_label and len(row) > 3:
                value = row[3]
                return float(value) if isinstance(value, (int, float)) else None
    return None


def parse_fin_health_excel(file_bytes: bytes, filename: str) -> dict[str, Any]:
    """Returns the evidence dict fin_health's rules score against. Every
    value here is read directly off a real cell — nothing is computed or
    guessed beyond simple text matching (e.g. the EPFO late-payment flag).
    """
    sheets = _read_rows(file_bytes, filename)

    about = _property_value_dict(sheets.get("About the Company", []))
    gst = _first_data_row_dict(sheets.get("GST", []))
    epfo = _first_data_row_dict(sheets.get("EPFO Establishments", []))
    financials = sheets.get("Standalone Financial Data", [])

    epfo_flags = str(epfo.get("FLAGS") or "")

    paid_up_capital_cr = about.get("Paid Up Capital (Crore)")
    total_secured_charges_cr = about.get("Sum of Charges (Crore)")
    charges_to_capital_ratio = (
        round(total_secured_charges_cr / paid_up_capital_cr, 2)
        if isinstance(total_secured_charges_cr, (int, float)) and isinstance(paid_up_capital_cr, (int, float))
        and paid_up_capital_cr != 0
        else None
    )

    return {
        "business_name": about.get("Legal Name"),
        "company_status": about.get("Company Status"),
        "paid_up_capital_cr": paid_up_capital_cr,
        "total_secured_charges_cr": total_secured_charges_cr,
        "charges_to_capital_ratio": charges_to_capital_ratio,
        "gst_status": gst.get("GSTIN STATUS"),
        "epfo_late_payment_flag": "payment after due date" in epfo_flags.lower(),
        "revenue_growth_pct": _latest_year_ratio(financials, "RATIOS - AOC-4", "Revenue Growth (%)"),
        "ebitda_margin_pct": _latest_year_ratio(financials, "RATIOS - AOC-4", "EBITDA Margin (%)"),
        "net_margin_pct": _latest_year_ratio(financials, "RATIOS - AOC-4", "Net Margin (%)"),
        "debt_to_equity": _latest_year_ratio(financials, "RATIOS - AOC-4", "Debt / Equity"),
        "interest_coverage_ratio": _latest_year_ratio(financials, "RATIOS - AOC-4", "Interest Coverage Ratio"),
        "current_ratio": _latest_year_ratio(financials, "RATIOS - AOC-4", "Current Ratio"),
    }
