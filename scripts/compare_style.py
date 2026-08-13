"""A/B the vernacular style layer and write the result to Excel.

Answers every question twice in one process -- once with style retrieval
forced off, once with it on -- and writes both answers side by side so
someone who reads Hindi can judge whether the register actually improved.

Run in-process rather than over HTTP on purpose: toggling style through the
API would mean restarting the backend between every pair, and the two runs
would then be minutes apart on a host whose load varies. Here the only
deliberate difference between A and B is whether style examples were in the
prompt.

Two columns matter more than the Hindi:

**tools_match** -- the two runs each do their own tool loop, so they can
diverge. If they did, the answers differ for reasons that have nothing to do
with style, and that row cannot be read as a style comparison.

**numbers_kept** -- every figure in the unstyled answer, checked against the
styled one. This is the whole safety question in one column: style is only
allowed to change wording, and a number that changes between A and B means
it did something it must not.

Usage:
    OLLAMA_HOST=... python scripts/compare_style.py
    OLLAMA_HOST=... python scripts/compare_style.py --questions my_prompts.json
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import time
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

import capabilities_impl  # noqa: E402,F401 - registers the capabilities
from agent_platform.runtime import chat  # noqa: E402
from capabilities_impl import style_examples  # noqa: E402

DEFAULT_QUESTIONS = [
    # Demo surface -- these must not regress.
    "मेरी बेटी 5 साल की है, उसके लिए कौन सी सरकारी योजना अच्छी है?",
    "1 लाख रुपये 1 साल की FD में रखूं तो मैच्योरिटी पर कितना मिलेगा?",
    "SBI में सेविंग्स अकाउंट पर अभी कितना ब्याज मिलता है?",
    "मैं महीने के 15000 कमाता हूँ, बैंक से सस्ता बीमा कौन सा मिलेगा?",
    # Where Hindi went wrong before -- scheme names and the age band.
    "PPF क्या है और इसमें कितना ब्याज मिलता है?",
    "मैं 62 साल का हूँ, हर महीने इनकम के लिए कौन सा विकल्प अच्छा है?",
    # Document-grounded, the path that used to answer from memory.
    "SBI में बचत खाता खोलने के लिए कौन से दस्तावेज चाहिए?",
    "अगर बैंक डूब जाए तो मेरा कितना पैसा सुरक्षित है?",
    "जीरो बैलेंस खाता क्या होता है और कौन खोल सकता है?",
    # Register stress: casual, conversational, no clean keyword.
    "भाई मेरी सैलरी 30 हज़ार है, पैसे कहाँ लगाऊं?",
    "होम लोन पर अभी कितना ब्याज लग रहा है?",
    "बुढ़ापे के लिए पेंशन का क्या इंतज़ाम करूं?",
]

# Money, rates, dates -- anything style must carry through untouched.
_FIGURE = re.compile(r"\d[\d,]*\.?\d*")


_SCALE = {"लाख": 100_000, "लाख": 100_000, "करोड़": 10_000_000, "करोड": 10_000_000,
          "हज़ार": 1_000, "हजार": 1_000}
_SCALED = re.compile(r"(\d[\d,]*\.?\d*)\s*(लाख|करोड़|करोड|हज़ार|हजार)")


def figures(text: str) -> set[float]:
    """Compared as numbers, not strings, and with Indian scale words resolved.

    Two false positives this has already produced, both of which made style
    look worse than it is:
      - "2.5%" rewritten as "2.50%" flagged as a lost rate.
      - "₹2,00,000" rewritten as "₹2 लाख" flagged as a lost amount, when it
        is the same sum said the way a person would say it -- which is
        precisely what the style layer is supposed to do.
    """
    text = text or ""
    values = set()
    consumed = []
    for amount, scale in _SCALED.findall(text):
        try:
            values.add(float(amount.replace(",", "")) * _SCALE[scale])
            consumed.append(f"{amount} {scale}")
        except ValueError:
            continue
    # Drop the scaled spans so "2 लाख" doesn't also register a bare 2.
    remainder = _SCALED.sub(" ", text)
    for match in _FIGURE.findall(remainder):
        try:
            values.add(float(match.replace(",", "")))
        except ValueError:
            continue
    return values


def ask(question: str, styled: bool, monkey) -> dict:
    """One turn. `styled=False` forces retrieval to return nothing, which is
    exactly the state the agent is in today with no index built."""
    original = style_examples.for_query
    if not styled:
        style_examples.for_query = lambda *_a, **_k: []
    started = time.time()
    try:
        result = chat.handle_chat_turn("finguru", None, question)
        reply = (result.reply or "").strip()
        error = ""
    except Exception as exc:                # noqa: BLE001 - a failed run is a result
        reply, error = "", f"{type(exc).__name__}: {exc}"
    finally:
        style_examples.for_query = original
    return {"reply": reply, "seconds": round(time.time() - started, 1), "error": error}


def tools_from_log(since_line: int) -> list[str]:
    log = REPO_ROOT / "logs" / "ollama_calls.jsonl"
    if not log.exists():
        return []
    names: list[str] = []
    for line in log.read_text(encoding="utf-8", errors="replace").splitlines()[since_line:]:
        if not line.strip():
            continue
        try:
            record = json.loads(line)
        except json.JSONDecodeError:
            continue
        names += re.findall(r"'name': '((?:docs|india|money|fx)\.[a-z_]+)'",
                            str(record.get("response")))
    return sorted(set(names))


def log_lines() -> int:
    log = REPO_ROOT / "logs" / "ollama_calls.jsonl"
    if not log.exists():
        return 0
    return len(log.read_text(encoding="utf-8", errors="replace").splitlines())


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--questions", help="JSON file with a 'questions' list")
    parser.add_argument("--out", default=str(REPO_ROOT / "docs" / "style-comparison.xlsx"))
    args = parser.parse_args()

    questions = DEFAULT_QUESTIONS
    if args.questions:
        loaded = json.loads(Path(args.questions).read_text(encoding="utf-8"))
        questions = loaded["questions"] if isinstance(loaded, dict) else loaded

    if not style_examples._load_index():
        print("WARNING: no style index built, so both columns will be identical.\n"
              "         Build one first: scripts/build_style_index.py\n", file=sys.stderr)

    rows = []
    for i, question in enumerate(questions, start=1):
        print(f"[{i}/{len(questions)}] {question[:60]}")

        mark = log_lines()
        without = ask(question, styled=False, monkey=None)
        tools_without = tools_from_log(mark)

        mark = log_lines()
        with_style = ask(question, styled=True, monkey=None)
        tools_with = tools_from_log(mark)

        missing = sorted(figures(without["reply"]) - figures(with_style["reply"]))
        missing = [f"{value:g}" for value in missing]

        rows.append({
            "question": question,
            "without_style": without["reply"] or without["error"],
            "with_style": with_style["reply"] or with_style["error"],
            "tools_without": ", ".join(tools_without) or "NONE",
            "tools_with": ", ".join(tools_with) or "NONE",
            "tools_match": "yes" if tools_without == tools_with else "NO — not comparable",
            "numbers_kept": "yes" if not missing else f"NO — lost {', '.join(missing[:4])}",
            "secs_without": without["seconds"],
            "secs_with": with_style["seconds"],
            "better?": "",
        })
        print(f"      tools {'match' if tools_without == tools_with else 'DIFFER'}"
              f" | numbers {'kept' if not missing else 'CHANGED'}"
              f" | {without['seconds']}s vs {with_style['seconds']}s")

    write_excel(rows, Path(args.out))
    kept = sum(1 for r in rows if r["numbers_kept"] == "yes")
    matched = sum(1 for r in rows if r["tools_match"] == "yes")
    print(f"\n{len(rows)} questions | tools matched {matched}/{len(rows)} "
          f"| numbers preserved {kept}/{len(rows)}")
    print(f"Wrote {args.out}")
    return 0


def write_excel(rows: list[dict], path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    book = Workbook()
    sheet = book.active
    sheet.title = "Style comparison"   # Excel rejects "/" in a sheet name

    headers = ["#", "Question (Hindi)", "WITHOUT style", "WITH style", "Tools (without)",
               "Tools (with)", "Same tools?", "Numbers preserved?", "Secs (without)",
               "Secs (with)", "Which reads better?"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    warn = PatternFill("solid", fgColor="FFC7CE")
    for i, row in enumerate(rows, start=1):
        sheet.append([i, row["question"], row["without_style"], row["with_style"],
                      row["tools_without"], row["tools_with"], row["tools_match"],
                      row["numbers_kept"], row["secs_without"], row["secs_with"], ""])
        line = sheet.max_row
        # Anything the reviewer must not skim past.
        for column in (7, 8):
            if sheet.cell(line, column).value != "yes":
                sheet.cell(line, column).fill = warn

    for column, width in zip("ABCDEFGHIJK", (4, 42, 68, 68, 24, 24, 14, 26, 12, 12, 20)):
        sheet.column_dimensions[column].width = width
    for line in range(2, sheet.max_row + 1):
        for column in range(1, len(headers) + 1):
            sheet.cell(line, column).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "C2"

    notes = book.create_sheet("How to read this")
    for line in [
        ["What this compares"],
        ["Each question was answered twice by the same agent in one run."],
        ["WITHOUT style = today's behaviour. WITH style = vernacular examples added to the prompt."],
        [""],
        ["The only question for a reviewer"],
        ["Does the WITH column read more like how a person would explain it — while saying the same thing?"],
        ["Fill in 'Which reads better?' with: without / with / same."],
        [""],
        ["Two columns to check first"],
        ["Same tools? — the runs each look things up separately. 'NO' means the answers differ for"],
        ["    reasons unrelated to style, so that row is not a fair comparison."],
        ["Numbers preserved? — every figure in the WITHOUT answer, checked against the WITH answer."],
        ["    'NO' is a defect, not a style choice. Style may change wording and nothing else."],
        [""],
        ["Red cells mean look here."],
    ]:
        notes.append(line)
    notes.column_dimensions["A"].width = 110
    notes["A1"].font = Font(bold=True, size=12)
    for heading in (5, 9):
        notes.cell(heading, 1).font = Font(bold=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)


if __name__ == "__main__":
    raise SystemExit(main())
